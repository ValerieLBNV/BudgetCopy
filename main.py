import openpyxl as xl
import shutil
import pandas
import pandas as pd
from openpyxl import Workbook, load_workbook
import os
from pathlib3x import Path


INPUT_PATH = "C:\\Users\\vval\\Timex Group\\Sangion, Martina - IKA\\Budget\\2022\\Budget Tracker\\Flash files per entity\\"
OUTPUT_PATH = "C:\\Users\\vval\\Timex Group\\Sangion, Martina - IKA\\Budget\\2022\\Budget Tracker\\Others\\Datasource Budget-Revenue tracker 2022.xlsx"
ARCHIVE_FOLDER_PATH = "C:\\Users\\vval\\Timex Group\\Sangion, Martina - IKA\\Budget\\2022\\Budget Tracker\\Flash files per entity\\Archive monthly versions\\"
SHEET_NAME = 'All'

dictionary = {"100AMS_IntlKA_Weekly Flash Update": '100 AMS',
              "220UK_IntlKA_Weekly Flash Update": '220 UK',
              "570 980 Vert_IntlKA_Weekly Flash Update": '570 980 Vert',
              "720TSH_IntlKA_Weekly Flash Update": '720 TSH',
              "788HK_IntlKA_Weekly Flash Update": '788 HK',
              "320GER_IntlKA_Weekly Flash Update": '320 GER'}

keys = list(dictionary.keys())
values = list(dictionary.values())


def get_version_cell(sheet):
    for row in sheet:
        for cell in row:
            value = str(cell.value)
            if value.lower() == "version":
                return cell
    raise ValueError('There is no specified cell')


def find_starting_cell(output_version_cell, output_worksheet, output_version_column, version):
    X = range(output_version_cell.row + 1, output_worksheet.max_row + 1)
    chosen_cell = ''
    for number in range(output_version_cell.row + 1,
                        output_worksheet.max_row + 1):  # row+1: excl.header, max+1 incl last max row
        chosen_cell = output_worksheet.cell(number, output_version_column)
        # case 1 where we found a cell with the same version
        if chosen_cell.value == int(version):
            overwrite = ""
            while overwrite.lower() != "yes" and overwrite.lower() != "no":
                overwrite = input(
                    "Version " + version + " in " + str(output_worksheet) + " already exists. Overwrite? Yes/no: ")
            return chosen_cell, overwrite  # return location chosen cell and overwrite "yes"
    # case 2 where we did not find a cell with the same version, and we return the last cell in the column
    overwrite = "yes"
    return output_worksheet.cell(chosen_cell.row + 1,
                                 chosen_cell.column), overwrite  # return location chosen cell and overwrite "no"


def copy_data(version, chosen_files):
    output_workbook = xl.load_workbook(OUTPUT_PATH)
    # loop every entity number chosen through
    for entity_number in chosen_files:
        # loading from input excel
        input_file_name = keys[int(entity_number) - 1] + ".xlsx"
        input_workbook = xl.load_workbook(INPUT_PATH + input_file_name, data_only=True)
        input_worksheet = input_workbook[SHEET_NAME]
        input_version_cell = get_version_cell(input_worksheet)
        input_version_cell_below = input_worksheet.cell(input_version_cell.row + 1,
                                                        input_version_cell.column)  # eliminate header row
        max_cell = input_worksheet.cell(input_worksheet.max_row, input_worksheet.max_column)
        cell_range = input_worksheet[input_version_cell_below.coordinate:max_cell.coordinate]
        # find version cell in output to get the version column
        output_sheet_name = values[int(entity_number) - 1]
        output_worksheet = output_workbook[output_sheet_name]
        output_version_cell = get_version_cell(output_worksheet)
        output_version_column = output_version_cell.column
        # return values from def find_starting_cell
        chosen_cell, overwrite = find_starting_cell(output_version_cell, output_worksheet, output_version_column,
                                                    version)
        # perform copy paste data range
        if overwrite.lower() == "no":
            print("Skipping")
            continue
        elif overwrite.lower() == "yes":
            # Paste cell range
            row_counter = 0
            for row in cell_range:
                cell_counter = 0
                for cell in row:
                    if cell_counter == 0:  # overwrite all value of version cells to indicated version
                        output_worksheet.cell(chosen_cell.row + row_counter, chosen_cell.column).value = int(version)
                    else:  # paste the rest of cell range
                        output_worksheet.cell(chosen_cell.row + row_counter,
                                              chosen_cell.column + cell_counter).value = cell.value
                    cell_counter += 1
                row_counter += 1
            print("Data copied from "+ input_file_name)
    output_workbook.save(OUTPUT_PATH)


# TODO: delete previous data of measure YTG (or find how in DAX)
def archive_action(input_file_name, version, new_file_name):
    # TODO: specify input file exclude sheet Overview
    #input_workbook = xl.load_workbook(INPUT_PATH + input_file_name + ".xlsx")
    #input_worksheet = input_workbook[SHEET_NAME]

    #output_workbook = Workbook()
    #output_worksheet = output_workbook.active
    #output_worksheet = SHEET_NAME
    #output_workbook.save(filename=ARCHIVE_FOLDER_PATH+new_file_name)
    #output_worksheet = output_workbook.create_sheet(SHEET_NAME)
    #for row in input_worksheet:
     #   for cell in row:
    #        output_worksheet[cell.coordinate].value = cell.value

    archive_file_path = shutil.copyfile(INPUT_PATH + input_file_name + ".xlsx", ARCHIVE_FOLDER_PATH + new_file_name)
    archive_file = xl.load_workbook(archive_file_path)
    #PANDAS archive_file = pandas.read_excel(io= archive_file_path, sheet_name=SHEET_NAME)
    consolidated_sheet = archive_file[SHEET_NAME]
    #PANDAS archive_file.at[5:,1] = int(version)
    #PANDAS print(archive_file.head(10))
    consolidated_sheet['A1'].value = int(version)
    #output_worksheet['A1'].value = int(version)
    archive_file.save(archive_file_path)
    #output_workbook.save()
    print("Archived " + input_file_name)


def archive(version, chosen_files):
    for entity_number in chosen_files:
        input_file_name = keys[int(entity_number) - 1]
        new_file_name = input_file_name + " " + version + ".xlsx"
        if Path(ARCHIVE_FOLDER_PATH + new_file_name).is_file():
            overwrite = ""
            while overwrite.lower() != "yes" and overwrite.lower() != "no":
                overwrite = input(new_file_name + " already exists. Overwrite? Yes/no: ")
                if overwrite.lower() == "yes":
                    archive_action(input_file_name, version, new_file_name)
                elif overwrite.lower() == "no":
                    print('Skipping')
                    continue
        else:
            archive_action(input_file_name, version, new_file_name)


def choose_files(type, version):
    print("Here are the files we have: ")
    a = 1
    for entity in dictionary:
        print("\t" + entity + ": " + str(a))
        a += 1
    choice = input("Which file(s) do you want to work with: ")
    files_chosen = []
    for character in choice:
        if character.isdigit():
            files_chosen.append(character)

    if type == "1":
        type_choice = "archive file"
    else:
        type_choice = "copy data"

    action = ""
    while action.lower() != "yes" and action.lower() != "no":
        print("Run " + type_choice + " version " + version + " for: " + str(files_chosen) + " proceed? yes/no")
        action = input()
    if action.lower() == "yes":
        return files_chosen
    else:
        choose_files(type, version)


def main():
    type = input("Do you want to archive file or copy data. Type 1 for archive, 2 for copy data: ")
    version = input("Specify version number: ")

    if type == "1":
        chosen_files = choose_files(type, version)
        archive(version, chosen_files)

    elif type == "2":
        chosen_files = choose_files(type, version)
        copy_data(version, chosen_files)

    else:
        main()


# main method
if __name__ == '__main__':
    main()
